# openpyxl - ler e alterar dados de uma planilha
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carrega arquivo
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome da planilha
sheetname = 'Minha Planilha'

# Ativando planilha
worksheet: Worksheet = workbook[sheetname]  # type:ignore

# Exibindo valores da planilha
row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23)  # type:ignore
    print()

# Editando valor de célula
# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)
